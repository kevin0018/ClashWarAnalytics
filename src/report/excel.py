import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from zipfile import BadZipFile

class ExcelReporter:
    @staticmethod
    def generate(data: list, clan_name: str, filename: str = "CWL_History.xlsx"):
        """
        Generates an Excel report with professional formatting.
        Filters out players with 0 attacks and 0 defenses.
        """
        # Convert data objects to DataFrame
        df = pd.DataFrame([vars(d) for d in data])
        
        # Filter out players with 0 attacks AND 0 defenses (non-participants)
        df = df[(df['attacks_used'] > 0) | (df['defense_count'] > 0)]
        
        if df.empty:
            print(f"[WARN] No players with participation found. Skipping Excel generation.")
            return
        
        # Add net_balance as a column (it's a property, so we calculate it)
        df['net_balance'] = df['stars_earned'] - df['stars_conceded']
        
        # Column mapping (Spanish headers for user-facing report)
        columns_map = {
            'name': 'Jugador',
            'town_hall_level': 'TH',
            'attacks_used': 'Ataques',
            'three_star_count': '3 Estrellas',
            'two_star_count': '2 Estrellas',
            'one_star_count': '1 Estrella',
            'zero_star_count': '0 Estrellas',
            'stars_earned': 'Estrellas Atq.',
            'avg_destruction': 'Destr %',
            'avg_stars_attack': 'Prom. Atq',
            'defense_count': 'Defensas',
            'stars_conceded': 'Estrellas Def.',
            'avg_stars_defense': 'Prom. Def',
            'avg_destruction_defense': '% Dest. Recib.',
            'net_balance': 'Balance Neto'
        }
        
        # Select and rename columns based on model fields
        cols_to_use = [c for c in columns_map.keys() if c in df.columns]
        df = df[cols_to_use].rename(columns=columns_map)
        
        # Reorder columns to match desired layout
        column_order = [
            'Jugador',
            'TH',
            'Ataques',
            '3 Estrellas',
            '2 Estrellas',
            '1 Estrella',
            '0 Estrellas',
            'Estrellas Atq.',
            'Destr %',
            'Prom. Atq',
            'Defensas',
            'Estrellas Def.',
            'Prom. Def',
            '% Dest. Recib.',
            'Balance Neto'
        ]
        
        # Ensure all columns exist and reorder
        df = df[[col for col in column_order if col in df.columns]]
        
        # Convert percentage columns from 0-100 scale to 0-1 scale for Excel format
        # Robust conversion: ensure values are numeric before division
        if 'Destr %' in df.columns:
            df['Destr %'] = pd.to_numeric(df['Destr %'], errors='coerce')
            df['Destr %'] = df['Destr %'].apply(lambda x: x / 100.0 if pd.notna(x) and isinstance(x, (int, float)) else x)
        if '% Dest. Recib.' in df.columns:
            df['% Dest. Recib.'] = pd.to_numeric(df['% Dest. Recib.'], errors='coerce')
            df['% Dest. Recib.'] = df['% Dest. Recib.'].apply(lambda x: x / 100.0 if pd.notna(x) and isinstance(x, (int, float)) else x)

        # Generate sheet name dynamically: "ClanName Jan 2026"
        safe_clan_name = "".join([c for c in clan_name if c.isalnum() or c in (' ', '_')]).strip()
        month_str = pd.Timestamp.now().strftime("%b %Y")
        sheet_name = f"{safe_clan_name} {month_str}"[:31]

        # File mode determination
        # Check for corruption and determine if we should append or write new
        mode = 'w'
        if os.path.exists(filename):
            try:
                # Test if file is valid
                load_workbook(filename).close()
                mode = 'a'
            except (BadZipFile, KeyError, ValueError):
                print(f"[WARN] The file {filename} was corrupted. A new one will be created.")
                mode = 'w'

        # Writing data with proper ExcelWriter mode
        if mode == 'a':
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace')
        else:
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w')

        with writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet objects for styling
            ws = writer.sheets[sheet_name]
            
            # Calculate dimensions
            last_row = len(df) + 1
            last_col = len(df.columns)
            last_col_letter = get_column_letter(last_col)
            
            # Create Excel Table with professional style
            ref = f"A1:{last_col_letter}{last_row}"
            
            # Table name must be unique and alphanumeric
            sanitized_table_name = f"CWL_{safe_clan_name.replace(' ', '')}_{month_str.replace(' ', '')}"
            sanitized_table_name = "".join([c for c in sanitized_table_name if c.isalnum()])
            
            tab = Table(displayName=sanitized_table_name, ref=ref)
            
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Manual formatting for headers (bold and fill)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Apply formatting to data cells
            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            
            # Apply number formats for percentage and decimal columns
            percentage_columns = ['Destr %', '% Dest. Recib.']
            decimal_columns = ['Prom. Atq', 'Prom. Def']
            
            # Create a mapping of column names to their indices
            col_name_to_idx = {name: idx for idx, name in enumerate(column_order, start=1) if name in df.columns}
            
            # Apply percentage format (0.00% shows as percentage with 2 decimals)
            for col_name in percentage_columns:
                if col_name in col_name_to_idx:
                    col_letter = get_column_letter(col_name_to_idx[col_name])
                    for row_idx in range(2, last_row + 1):
                        cell = ws[f"{col_letter}{row_idx}"]
                        if cell.value is not None:
                            cell.number_format = '0.00%'
            
            # Apply decimal format for average columns (2 decimal places)
            for col_name in decimal_columns:
                if col_name in col_name_to_idx:
                    col_letter = get_column_letter(col_name_to_idx[col_name])
                    for row_idx in range(2, last_row + 1):
                        cell = ws[f"{col_letter}{row_idx}"]
                        if cell.value is not None:
                            cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Conditional Formatting (ColorScale) for "Balance Neto" column
            net_col_letter = None
            target_header = "Balance Neto"
            
            for cell in ws[1]:
                if cell.value == target_header:
                    net_col_letter = cell.column_letter
                    break
            
            if net_col_letter:
                rule = ColorScaleRule(
                    start_type='num', start_value=-7, start_color='FF9999',  # Red
                    mid_type='num', mid_value=0, mid_color='FFFFFF',          # White
                    end_type='num', end_value=7, end_color='99FF99'           # Green
                )
                ws.conditional_formatting.add(f"{net_col_letter}2:{net_col_letter}{last_row}", rule)

        print(f"[EXCEL] Report generated successfully in the sheet: {sheet_name}")
